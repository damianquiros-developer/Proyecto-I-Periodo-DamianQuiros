"""
Agenda web de contactos.
Flask funciona como el puente entre las paginas HTML y Python.
El Excel es la libreta donde se guardan usuarios, contactos y actividad.
"""

from flask import Flask, render_template, request, jsonify, session
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import re
import uuid
import datetime
import json
import io
import csv as modulo_csv
import os
from pathlib import Path
import openpyxl
from openpyxl import Workbook

# Aqui empieza todo.

app = Flask(__name__,
            template_folder='templates',
            static_folder='static',
            static_url_path='/static')

app.secret_key = os.environ.get("SECRET_KEY", "fiducci-secret-key-2024")
app.url_map.strict_slashes = False
CORS(app, supports_credentials=True)

# El Excel es donde se guardan los datos.

RUTA_BD = Path(__file__).parent / "contactos.xlsx"

COLS_USUARIOS      = ["id", "nombre", "apellido", "usuario", "correo",
                      "contrasena", "rol", "creado_en", "telefono", "cargo", "empresa"]
COLS_CONTACTOS     = ["Nombre", "Apellido", "Teléfono", "Correo", "Dirección",
                      "Categoría", "Favorito"]
COLS_CONTACTOS_META = ["id", "usuario_id", "empresa", "cargo", "creado_en"]
COLS_ACTIVIDAD     = ["id", "tipo", "descripcion", "fecha", "contacto_id", "usuario_id"]
COLS_INTERACCIONES = ["id", "contacto_id", "tipo", "descripcion", "fecha", "creado_por", "usuario_id"]
COLS_RECORDATORIOS = ["id", "contacto_id", "titulo", "fecha_limite", "completado", "creado_en", "usuario_id"]

# Esto sirve como traductor entre la pagina y el Excel.
_MAPA_API_A_EXCEL = {
    "nombre":    "Nombre",
    "apellido":  "Apellido",
    "telefono":  "Teléfono",
    "email":     "Correo",
    "notas":     "Dirección",
    "categoria": "Categoría",
}

_MAPA_API_A_META = {
    "empresa":   "empresa",
    "cargo":     "cargo",
}


def _ahora():
    return datetime.datetime.now().isoformat()


def _nuevo_id():
    return str(uuid.uuid4())


# Acomodar datos entre el Excel y la pagina.

def _normalizar_contacto(c_excel, meta=None):
    """Acomoda un contacto para que la pagina lo pueda usar."""
    meta = meta or {}
    return {
        "id":          meta.get("id") or c_excel.get("id") or "",
        "nombre":      c_excel.get("Nombre") or "",
        "apellido":    c_excel.get("Apellido") or "",
        "email":       c_excel.get("Correo") or "",
        "telefono":    c_excel.get("Teléfono") or "",
        "notas":       c_excel.get("Dirección") or "",
        "categoria":   c_excel.get("Categoría") or "",
        "es_favorito": bool(c_excel.get("Favorito")),
        "empresa":     meta.get("empresa") or c_excel.get("empresa") or "",
        "cargo":       meta.get("cargo") or c_excel.get("cargo") or "",
        "creado_en":   meta.get("creado_en") or c_excel.get("creado_en") or "",
        "usuario_id":  meta.get("usuario_id") or c_excel.get("usuario_id") or "",
    }


def _fila_excel_contacto(datos):
    """Arma una fila lista para guardarla en Contactos."""
    return [
        (datos.get("nombre")    or "").strip(),
        (datos.get("apellido")  or "").strip(),
        (datos.get("telefono")  or "").strip(),
        (datos.get("email")     or "").strip().lower(),
        (datos.get("notas")     or "").strip(),
        datos.get("categoria")  or "",
        False,
    ]


def _fila_meta_contacto(datos, id_c, id_usuario, creado_en=None):
    """Guarda datos extra sin meterlos en la hoja que revisa la rubrica."""
    return [
        id_c,
        id_usuario,
        (datos.get("empresa")   or "").strip(),
        (datos.get("cargo")     or "").strip(),
        creado_en or _ahora(),
    ]


# Abrir o crear el Excel.

def _iniciar_libro():
    """Crea el Excel desde cero con el usuario admin y contactos de prueba."""
    libro = Workbook()

    hoja_usuarios       = libro.active
    hoja_usuarios.title = "Usuarios"
    hoja_usuarios.append(COLS_USUARIOS)
    id_admin = _nuevo_id()
    hoja_usuarios.append([
        id_admin, "Administrador", "Sistema", "admin",
        "admin@fiducci.com", generate_password_hash("1234"), "admin", _ahora(),
        "", "", "FIDUCCI"
    ])

    hoja_contactos = libro.create_sheet("Contactos")
    hoja_contactos.append(COLS_CONTACTOS)
    hoja_meta = libro.create_sheet("ContactosMeta")
    hoja_meta.append(COLS_CONTACTOS_META)
    ahora = _ahora()
    for fila in [
        ["María",    "González",  "88881111", "maria.gonzalez@empresa.com",
         "Av. Central, San José",      "VIP",       False, _nuevo_id(), "TechCorp",           "Gerente"],
        ["Carlos",   "Rodríguez", "77772222", "carlos.r@finanzas.com",
         "Calle 5, Heredia",            "Familia",   False, _nuevo_id(), "Finanzas SA",        "Director"],
        ["Ana",      "Vargas",    "66663333", "ana.v@socios.com",
         "Residencial Los Cedros",      "Amigos",    True,  _nuevo_id(), "Vargas & Asociados", "Socia"],
        ["Luis",     "Mora",      "55554444", "luis.mora@prospecto.com",
         "San Pedro de Montes de Oca",  "Trabajo",   False, _nuevo_id(), "Mora Imports",       "Gerente"],
        ["Sofía",    "Jiménez",   "44445555", "sofia.j@correo.com",
         "Avenida 10, Cartago",         "Otro",      True,  _nuevo_id(), "Instituto CRA",      "Profesora"],
    ]:
        hoja_contactos.append(fila[:7])
        hoja_meta.append([fila[7], id_admin, fila[8], fila[9], ahora])

    hoja_actividad = libro.create_sheet("Actividad")
    hoja_actividad.append(COLS_ACTIVIDAD)
    for tipo, desc in [
        ("creado",   "Contacto María González agregado"),
        ("creado",   "Contacto Ana Vargas agregado"),
        ("favorito", "Ana Vargas marcada como favorita"),
        ("creado",   "Contacto Sofía Jiménez agregado"),
        ("favorito", "Sofía Jiménez marcada como favorita"),
    ]:
        hoja_actividad.append([_nuevo_id(), tipo, desc, ahora, "", id_admin])

    libro.save(RUTA_BD)
    return libro


def _cargar_libro():
    """Abre el Excel. Si no existe, lo crea como cuando se abre libreta nueva."""
    if not RUTA_BD.exists():
        return _iniciar_libro()
    libro = openpyxl.load_workbook(RUTA_BD)
    return _migrar_si_necesario(libro)


def _encabezados(hoja):
    return [hoja.cell(1, j).value for j in range(1, hoja.max_column + 1)]


def _asegurar_contactos_meta(libro):
    """Revisa que exista la hoja donde van datos extra de cada contacto."""
    if "ContactosMeta" not in libro.sheetnames:
        hoja_meta = libro.create_sheet("ContactosMeta")
        hoja_meta.append(COLS_CONTACTOS_META)
    return libro["ContactosMeta"]


def _reconstruir_contactos_y_meta(libro, registros, id_admin):
    """Acomoda Contactos con las 7 columnas pedidas y guarda lo extra aparte."""
    hoja_c = libro["Contactos"]
    hoja_c.delete_rows(1, hoja_c.max_row)
    hoja_c.append(COLS_CONTACTOS)

    hoja_m = _asegurar_contactos_meta(libro)
    hoja_m.delete_rows(1, hoja_m.max_row)
    hoja_m.append(COLS_CONTACTOS_META)

    for d in registros:
        contacto = [
            d.get("Nombre") or d.get("nombre") or "",
            d.get("Apellido") or d.get("apellido") or "",
            d.get("Teléfono") or d.get("telefono") or "",
            d.get("Correo") or d.get("email") or "",
            d.get("Dirección") or d.get("notas") or "",
            d.get("Categoría") or d.get("categoria") or "",
            bool(d.get("Favorito") or d.get("es_favorito")),
        ]
        meta = [
            d.get("id") or _nuevo_id(),
            d.get("usuario_id") or id_admin,
            d.get("empresa") or "",
            d.get("cargo") or "",
            d.get("creado_en") or _ahora(),
        ]
        hoja_c.append(contacto)
        hoja_m.append(meta)


def _migrar_si_necesario(libro):
    """Acomoda un Excel viejo para que sirva con esta version."""
    cambios = False

    # Si faltan datos de perfil, se agregan sin tocar lo que ya habia.
    hoja_u = libro["Usuarios"]
    enc_u  = _encabezados(hoja_u)
    for col_nueva in ["telefono", "cargo", "empresa"]:
        if col_nueva not in enc_u:
            pos = hoja_u.max_column + 1
            hoja_u.cell(1, pos).value = col_nueva
            for i in range(2, hoja_u.max_row + 1):
                hoja_u.cell(i, pos).value = ""
            enc_u.append(col_nueva)
            cambios = True

    # Soporta nombres de columnas viejas.
    if "name" in enc_u and "nombre" not in enc_u:
        for i in range(1, hoja_u.max_row + 1):
            fila = {enc_u[j - 1]: hoja_u.cell(i, j).value for j in range(1, hoja_u.max_column + 1)}
            for nuevo, viejo in [("nombre", "name"), ("apellido", "lastname"),
                                  ("usuario", "username"), ("correo", "email"),
                                  ("contrasena", "password_hash"), ("rol", "role")]:
                if viejo in enc_u:
                    hoja_u.cell(i, enc_u.index(viejo) + 1).value = fila.get(viejo)
        cambios = True

    # La hoja Contactos queda igual a la rubrica y lo extra se guarda aparte.
    hoja_c = libro["Contactos"]
    enc_c  = _encabezados(hoja_c)
    hoja_m = libro["ContactosMeta"] if "ContactosMeta" in libro.sheetnames else None
    enc_m  = _encabezados(hoja_m) if hoja_m else []
    meta_ok = bool(hoja_m) and enc_m == COLS_CONTACTOS_META and hoja_m.max_row == hoja_c.max_row

    if enc_c != COLS_CONTACTOS or not meta_ok:
        id_admin = _primer_admin(libro)
        registros = []
        for i in range(2, hoja_c.max_row + 1):
            fila_vals = [hoja_c.cell(i, j).value for j in range(1, hoja_c.max_column + 1)]
            if not any(v is not None for v in fila_vals):
                continue
            registro = dict(zip(enc_c, fila_vals))
            if hoja_m and i <= hoja_m.max_row:
                meta_vals = [hoja_m.cell(i, j).value for j in range(1, hoja_m.max_column + 1)]
                for clave, valor in dict(zip(enc_m, meta_vals)).items():
                    if valor not in (None, "") or clave not in registro:
                        registro[clave] = valor
            registros.append(registro)

        _reconstruir_contactos_y_meta(libro, registros, id_admin)
        cambios = True

    # La actividad necesita saber de que usuario es cada accion.
    if "Actividad" not in libro.sheetnames:
        hoja_a = libro.create_sheet("Actividad")
        hoja_a.append(COLS_ACTIVIDAD)
        cambios = True
    hoja_a = libro["Actividad"]
    enc_a  = _encabezados(hoja_a)
    if "usuario_id" not in enc_a:
        id_admin = _primer_admin(libro)
        pos = hoja_a.max_column + 1
        hoja_a.cell(1, pos).value = "usuario_id"
        for i in range(2, hoja_a.max_row + 1):
            if any(hoja_a.cell(i, j).value for j in range(1, pos)):
                hoja_a.cell(i, pos).value = id_admin
        cambios = True

    if cambios:
        _guardar_libro(libro)
    return libro


def _primer_admin(libro):
    """Busca el primer admin para usarlo cuando hay datos viejos sin usuario."""
    hoja_u = libro["Usuarios"]
    enc_u  = [hoja_u.cell(1, j).value for j in range(1, hoja_u.max_column + 1)]
    rol_col = "rol" if "rol" in enc_u else "role"
    id_col  = "id"
    for fila in hoja_u.iter_rows(min_row=2, values_only=True):
        d = dict(zip(enc_u, fila))
        if d.get(rol_col) == "admin":
            return d.get(id_col, "")
    for fila in hoja_u.iter_rows(min_row=2, values_only=True):
        d = dict(zip(enc_u, fila))
        return d.get(id_col, "")
    return ""


def _guardar_libro(libro):
    libro.save(RUTA_BD)


def _hoja_a_dicts(hoja, cols):
    """Pasa las filas del Excel a una forma mas facil de leer."""
    resultado = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in fila):
            resultado.append(dict(zip(cols, fila)))
    return resultado


def _buscar_fila(hoja, col_idx, valor):
    """Busca una fila por una columna, como buscar una cedula en una lista."""
    for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        if fila[col_idx] == valor:
            return i
    return None


def _registrar(libro, tipo, descripcion, id_contacto="", id_usuario=""):
    """Guarda una accion en actividad reciente."""
    libro["Actividad"].append([_nuevo_id(), tipo, descripcion, _ahora(),
                               id_contacto or "", id_usuario or ""])


# Usuarios.

def _buscar_usuario(identificador):
    """Busca un usuario por usuario o correo."""
    libro    = _cargar_libro()
    hoja_u   = libro["Usuarios"]
    enc_u    = [hoja_u.cell(1, j).value for j in range(1, hoja_u.max_column + 1)]
    usuarios = _hoja_a_dicts(hoja_u, enc_u)
    id_min   = identificador.lower()
    for u in usuarios:
        usu_col  = u.get("usuario") or u.get("username") or ""
        cor_col  = u.get("correo")  or u.get("email")    or ""
        cont_col = u.get("contrasena") or u.get("password_hash") or ""
        if usu_col.lower() == id_min or cor_col.lower() == id_min:
            return {
                "id":            u.get("id", ""),
                "name":          u.get("nombre") or u.get("name") or "",
                "lastname":      u.get("apellido") or u.get("lastname") or "",
                "username":      usu_col,
                "email":         cor_col,
                "password_hash": cont_col,
                "role":          u.get("rol") or u.get("role") or "user",
                "telefono":      u.get("telefono") or "",
                "cargo":         u.get("cargo") or "",
                "empresa":       u.get("empresa") or "",
            }
    return None


def _crear_usuario(nombre, apellido, correo, contrasena):
    """Crea un usuario nuevo si el correo todavia no existe."""
    libro    = _cargar_libro()
    hoja_u   = libro["Usuarios"]
    enc_u    = [hoja_u.cell(1, j).value for j in range(1, hoja_u.max_column + 1)]
    usuarios = _hoja_a_dicts(hoja_u, enc_u)
    cor_col  = "correo" if "correo" in enc_u else "email"

    if any((u.get(cor_col) or "").lower() == correo for u in usuarios):
        return None

    base     = nombre.lower().replace(" ", "_")
    usu_col  = "usuario" if "usuario" in enc_u else "username"
    existentes = {u.get(usu_col, "") for u in usuarios}
    nom_usu  = base
    n = 1
    while nom_usu in existentes:
        nom_usu = f"{base}{n}"; n += 1

    nuevo = {
        "id":          _nuevo_id(),
        "nombre":      nombre,
        "apellido":    apellido,
        "usuario":     nom_usu,
        "correo":      correo,
        "contrasena":  generate_password_hash(contrasena),
        "rol":         "user",
        "creado_en":   _ahora(),
        "telefono":    "",
        "cargo":       "",
        "empresa":     "",
    }
    hoja_u.append([nuevo[c] for c in COLS_USUARIOS])
    _guardar_libro(libro)
    return nuevo


def _actualizar_usuario(id_usuario, datos):
    """Actualiza el perfil del usuario que esta conectado."""
    libro  = _cargar_libro()
    hoja_u = libro["Usuarios"]
    enc_u  = [hoja_u.cell(1, j).value for j in range(1, hoja_u.max_column + 1)]
    fila   = _buscar_fila(hoja_u, 0, id_usuario)
    if fila is None:
        return None

    def _asignar(col_esp, col_en, valor):
        col = col_esp if col_esp in enc_u else col_en
        if col in enc_u:
            hoja_u.cell(fila, enc_u.index(col) + 1).value = valor

    if "name"          in datos: _asignar("nombre",    "name",          datos["name"])
    if "lastname"      in datos: _asignar("apellido",  "lastname",      datos["lastname"])
    if "email"         in datos: _asignar("correo",    "email",         datos["email"])
    if "telefono"      in datos: _asignar("telefono",  "telefono",      datos["telefono"])
    if "cargo"         in datos: _asignar("cargo",     "cargo",         datos["cargo"])
    if "empresa"       in datos: _asignar("empresa",   "empresa",       datos["empresa"])
    if "password_hash" in datos: _asignar("contrasena","password_hash", datos["password_hash"])

    _guardar_libro(libro)

    return _buscar_usuario(
        hoja_u.cell(fila, enc_u.index("usuario" if "usuario" in enc_u else "username") + 1).value or ""
    )


# Revisar datos.

def _validar_telefono(telefono):
    """Revisa que el telefono tenga 8 digitos."""
    solo_digitos = re.sub(r'\D', '', telefono or "")
    return len(solo_digitos) == 8


def _validar_correo(correo):
    """Revisa que el correo tenga @ y dominio."""
    return bool(re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', correo or ""))


def _validar_contacto_api(datos):
    """Revisa el contacto antes de guardarlo en el Excel."""
    nombre    = (datos.get("nombre")   or "").strip()
    apellido  = (datos.get("apellido") or "").strip()
    telefono  = (datos.get("telefono") or "").strip()
    correo    = (datos.get("email")    or "").strip()
    categoria = (datos.get("categoria") or "").strip()

    if not nombre:
        return "El nombre es requerido"
    if not apellido:
        return "El apellido es requerido"
    if not telefono:
        return "El teléfono es requerido"
    if not _validar_telefono(telefono):
        return "El teléfono debe tener exactamente 8 dígitos"
    if not correo:
        return "El correo es requerido"
    if not _validar_correo(correo):
        return "El formato del correo no es válido (debe contener @ y un dominio)"
    if not categoria:
        return "La categoría es requerida"
    return None


# Contactos.

def _obtener_contactos(id_usuario):
    """Devuelve solo los contactos del usuario conectado."""
    libro  = _cargar_libro()
    hoja_c = libro["Contactos"]
    hoja_m = _asegurar_contactos_meta(libro)
    resultado = []
    for i in range(2, hoja_c.max_row + 1):
        c = {COLS_CONTACTOS[j - 1]: hoja_c.cell(i, j).value
             for j in range(1, len(COLS_CONTACTOS) + 1)}
        if not any(v is not None for v in c.values()):
            continue
        meta = {COLS_CONTACTOS_META[j - 1]: hoja_m.cell(i, j).value
                for j in range(1, len(COLS_CONTACTOS_META) + 1)}
        uid = meta.get("usuario_id") or ""
        if uid != id_usuario:
            continue
        resultado.append(_normalizar_contacto(c, meta))
    return resultado


def _crear_contacto(datos, id_usuario):
    """Guarda un contacto nuevo y deja registrada la accion."""
    id_c     = _nuevo_id()
    creado   = _ahora()
    fila     = _fila_excel_contacto(datos)
    fila_meta = _fila_meta_contacto(datos, id_c, id_usuario, creado)
    nuevo    = _normalizar_contacto(dict(zip(COLS_CONTACTOS, fila)),
                                    dict(zip(COLS_CONTACTOS_META, fila_meta)))

    libro  = _cargar_libro()
    hoja_m = _asegurar_contactos_meta(libro)
    libro["Contactos"].append(fila)
    hoja_m.append(fila_meta)
    _registrar(libro, "creado",
               f"Contacto {nuevo['nombre']} {nuevo['apellido']} agregado",
               id_c, id_usuario)
    _guardar_libro(libro)
    return nuevo


def _actualizar_contacto(id_contacto, datos, id_usuario):
    """Edita un contacto si pertenece al usuario conectado."""
    libro  = _cargar_libro()
    hoja_c = libro["Contactos"]
    hoja_m = _asegurar_contactos_meta(libro)
    idx_id = COLS_CONTACTOS_META.index("id")
    fila   = _buscar_fila(hoja_m, idx_id, id_contacto)
    if fila is None:
        return None

    idx_usu = COLS_CONTACTOS_META.index("usuario_id")
    if hoja_m.cell(fila, idx_usu + 1).value != id_usuario:
        return None

    # Primero se lee la fila completa, como tener la tarjeta en la mano.
    fila_dict = {COLS_CONTACTOS[j - 1]: hoja_c.cell(fila, j).value
                 for j in range(1, len(COLS_CONTACTOS) + 1)}
    meta_dict = {COLS_CONTACTOS_META[j - 1]: hoja_m.cell(fila, j).value
                 for j in range(1, len(COLS_CONTACTOS_META) + 1)}

    # Luego se cambian solo los campos que llegaron desde el formulario.
    for campo_api, col_excel in _MAPA_API_A_EXCEL.items():
        if campo_api in datos and col_excel in COLS_CONTACTOS:
            col_idx = COLS_CONTACTOS.index(col_excel) + 1
            hoja_c.cell(fila, col_idx).value = datos[campo_api]
            fila_dict[col_excel] = datos[campo_api]
    for campo_api, col_meta in _MAPA_API_A_META.items():
        if campo_api in datos and col_meta in COLS_CONTACTOS_META:
            col_idx = COLS_CONTACTOS_META.index(col_meta) + 1
            hoja_m.cell(fila, col_idx).value = datos[campo_api]
            meta_dict[col_meta] = datos[campo_api]

    _registrar(libro, "editado",
               f"Contacto {fila_dict.get('Nombre', '')} {fila_dict.get('Apellido', '')} actualizado",
               id_contacto, id_usuario)
    _guardar_libro(libro)
    return _normalizar_contacto(fila_dict, meta_dict)


def _eliminar_contacto(id_contacto, id_usuario):
    """Elimina un contacto y tambien su fila de datos internos."""
    libro  = _cargar_libro()
    hoja_c = libro["Contactos"]
    hoja_m = _asegurar_contactos_meta(libro)
    idx_id = COLS_CONTACTOS_META.index("id")
    fila   = _buscar_fila(hoja_m, idx_id, id_contacto)
    if fila is None:
        return False

    idx_usu = COLS_CONTACTOS_META.index("usuario_id")
    if hoja_m.cell(fila, idx_usu + 1).value != id_usuario:
        return False

    nombre   = hoja_c.cell(fila, COLS_CONTACTOS.index("Nombre")   + 1).value or ""
    apellido = hoja_c.cell(fila, COLS_CONTACTOS.index("Apellido") + 1).value or ""
    hoja_c.delete_rows(fila)
    hoja_m.delete_rows(fila)
    _registrar(libro, "eliminado",
               f"Contacto {nombre} {apellido} eliminado",
               id_contacto, id_usuario)
    _guardar_libro(libro)
    return True


def _alternar_favorito(id_contacto, id_usuario):
    """Prende o apaga la estrella de favorito."""
    libro  = _cargar_libro()
    hoja_c = libro["Contactos"]
    hoja_m = _asegurar_contactos_meta(libro)
    idx_id = COLS_CONTACTOS_META.index("id")
    fila   = _buscar_fila(hoja_m, idx_id, id_contacto)
    if fila is None:
        return None

    idx_usu = COLS_CONTACTOS_META.index("usuario_id")
    if hoja_m.cell(fila, idx_usu + 1).value != id_usuario:
        return None

    idx_fav    = COLS_CONTACTOS.index("Favorito") + 1
    nuevo_val  = not bool(hoja_c.cell(fila, idx_fav).value)
    hoja_c.cell(fila, idx_fav).value = nuevo_val

    nombre   = hoja_c.cell(fila, COLS_CONTACTOS.index("Nombre")   + 1).value or ""
    apellido = hoja_c.cell(fila, COLS_CONTACTOS.index("Apellido") + 1).value or ""
    accion   = "marcado como favorito" if nuevo_val else "quitado de favoritos"
    _registrar(libro, "favorito", f"{nombre} {apellido} {accion}", id_contacto, id_usuario)
    _guardar_libro(libro)
    return nuevo_val


# Actividad reciente.

def _asegurar_hoja(libro, nombre, cols):
    """Crea una hoja si todavia no existe en el Excel."""
    if nombre not in libro.sheetnames:
        hoja = libro.create_sheet(nombre)
        hoja.append(cols)
        _guardar_libro(libro)
    return libro[nombre]


def _obtener_actividad(id_usuario, limite=50):
    """Devuelve las acciones recientes, de la ultima hacia atras."""
    libro     = _cargar_libro()
    elementos = _hoja_a_dicts(libro["Actividad"], COLS_ACTIVIDAD)
    resultado = []
    for item in elementos:
        if item.get("usuario_id") != id_usuario:
            continue
        for k in list(item.keys()):
            if item[k] is None:
                item[k] = ""
        resultado.append(item)
    resultado.reverse()
    return resultado[:limite]


# Interacciones.

def _obtener_interacciones(id_contacto, id_usuario):
    libro = _cargar_libro()
    _asegurar_hoja(libro, "Interacciones", COLS_INTERACCIONES)
    elementos = _hoja_a_dicts(libro["Interacciones"], COLS_INTERACCIONES)
    resultado = []
    for item in elementos:
        if item.get("contacto_id") == id_contacto and item.get("usuario_id") == id_usuario:
            for k in list(item.keys()):
                if item[k] is None:
                    item[k] = ""
            resultado.append(item)
    resultado.reverse()
    return resultado


def _crear_interaccion(id_contacto, tipo, descripcion, creado_por, id_usuario):
    libro = _cargar_libro()
    _asegurar_hoja(libro, "Interacciones", COLS_INTERACCIONES)
    nueva = [_nuevo_id(), id_contacto, tipo, descripcion, _ahora(), creado_por, id_usuario]
    libro["Interacciones"].append(nueva)
    _guardar_libro(libro)
    return dict(zip(COLS_INTERACCIONES, nueva))


def _eliminar_interaccion(id_interaccion, id_usuario):
    libro = _cargar_libro()
    _asegurar_hoja(libro, "Interacciones", COLS_INTERACCIONES)
    hoja  = libro["Interacciones"]
    fila  = _buscar_fila(hoja, 0, id_interaccion)
    if fila is None:
        return False
    if hoja.cell(fila, COLS_INTERACCIONES.index("usuario_id") + 1).value != id_usuario:
        return False
    hoja.delete_rows(fila)
    _guardar_libro(libro)
    return True


# Recordatorios.

def _obtener_recordatorios(id_usuario):
    libro = _cargar_libro()
    _asegurar_hoja(libro, "Recordatorios", COLS_RECORDATORIOS)
    elementos = _hoja_a_dicts(libro["Recordatorios"], COLS_RECORDATORIOS)
    resultado = []
    for r in elementos:
        if r.get("usuario_id") != id_usuario:
            continue
        r["completado"] = bool(r.get("completado"))
        for k in list(r.keys()):
            if r[k] is None:
                r[k] = ""
        resultado.append(r)
    return resultado


def _obtener_recordatorios_contacto(id_contacto, id_usuario):
    return [r for r in _obtener_recordatorios(id_usuario)
            if r.get("contacto_id") == id_contacto]


def _crear_recordatorio(id_contacto, titulo, fecha_limite, id_usuario):
    libro = _cargar_libro()
    _asegurar_hoja(libro, "Recordatorios", COLS_RECORDATORIOS)
    nuevo = [_nuevo_id(), id_contacto, titulo, fecha_limite, False, _ahora(), id_usuario]
    libro["Recordatorios"].append(nuevo)
    _guardar_libro(libro)
    rec = dict(zip(COLS_RECORDATORIOS, nuevo))
    rec["completado"] = False
    return rec


def _completar_recordatorio(id_rec, id_usuario):
    libro = _cargar_libro()
    _asegurar_hoja(libro, "Recordatorios", COLS_RECORDATORIOS)
    hoja  = libro["Recordatorios"]
    fila  = _buscar_fila(hoja, 0, id_rec)
    if fila is None:
        return None
    if hoja.cell(fila, COLS_RECORDATORIOS.index("usuario_id") + 1).value != id_usuario:
        return None
    hoja.cell(fila, COLS_RECORDATORIOS.index("completado") + 1).value = True
    _guardar_libro(libro)
    return True


def _eliminar_recordatorio(id_rec, id_usuario):
    libro = _cargar_libro()
    _asegurar_hoja(libro, "Recordatorios", COLS_RECORDATORIOS)
    hoja  = libro["Recordatorios"]
    fila  = _buscar_fila(hoja, 0, id_rec)
    if fila is None:
        return False
    if hoja.cell(fila, COLS_RECORDATORIOS.index("usuario_id") + 1).value != id_usuario:
        return False
    hoja.delete_rows(fila)
    _guardar_libro(libro)
    return True


# Reporte.

def _generar_reporte(id_usuario):
    """Cuenta contactos, favoritos y categorias para mostrar el reporte."""
    contactos = _obtener_contactos(id_usuario)
    total     = len(contactos)
    favoritos = sum(1 for c in contactos if c.get("es_favorito"))
    por_categoria = {}
    for c in contactos:
        cat = c.get("categoria") or "Sin categoría"
        por_categoria[cat] = por_categoria.get(cat, 0) + 1
    return {
        "total":        total,
        "favoritos":    favoritos,
        "normales":     total - favoritos,
        "por_categoria": por_categoria,
    }


# Revisar si el usuario entro.

def _verificar_auth():
    """Evita que alguien vea datos sin entrar primero."""
    if not session.get("idUsuario"):
        nom = session.get("nombreUsuario")
        if not nom:
            return jsonify({"success": False, "message": "No autenticado"}), 401
        usuario = _buscar_usuario(nom)
        if not usuario:
            return jsonify({"success": False, "message": "No autenticado"}), 401
        session["idUsuario"] = usuario["id"]
    return None


# Paginas del sistema.

@app.route('/')
def pagina_inicio():
    return render_template('index.html')


@app.route('/favicon.ico')
def favicon():
    return "", 204


@app.route('/register')
def pagina_registro():
    return render_template('registroNuevo.html')


@app.route('/forgot-password')
def pagina_recuperar():
    return render_template('olvContrasena.html')


@app.route('/forgot-success')
def pagina_correo_enviado():
    return render_template('correoEnviado.html')


@app.route('/paginaPrincipal')
@app.route('/dashboard')
def pagina_principal():
    return render_template('paginaPrincipal.html')


@app.route('/perfil')
def pagina_perfil():
    return render_template('perfilUsuario.html')


@app.route('/favoritos')
def pagina_favoritos():
    return render_template('favoritos.html')


@app.route('/actividad')
def pagina_actividad():
    return render_template('actividadReciente.html')


@app.route('/contacto/nuevo')
def pagina_nuevo_contacto():
    return render_template('nuevoContacto.html')


@app.route('/contacto/detalle')
def pagina_detalle_contacto():
    return render_template('detalleContacto.html')


@app.route('/contacto/editar')
def pagina_editar_contacto():
    return render_template('editarContacto.html')


@app.route('/reporte')
def pagina_reporte():
    return render_template('reporte.html')


# Pedidos de entrada y registro.

@app.route("/api/login", methods=["POST"])
def api_iniciar_sesion():
    datos         = request.get_json(silent=True) or {}
    identificador = (datos.get("email") or "").strip()
    contrasena    = datos.get("password") or ""

    if not identificador or not contrasena:
        return jsonify({"success": False, "message": "Campos requeridos"}), 400

    usuario = _buscar_usuario(identificador)
    if not usuario or not check_password_hash(usuario["password_hash"], contrasena):
        return jsonify({"success": False, "message": "Credenciales incorrectas"}), 401

    session["idUsuario"]     = usuario["id"]
    session["nombreUsuario"] = usuario["username"]

    return jsonify({
        "success": True,
        "user": {
            "id":       usuario["id"],
            "name":     usuario["name"],
            "lastname": usuario["lastname"],
            "email":    usuario["email"],
            "role":     usuario["role"],
        }
    })


@app.route("/api/register", methods=["POST"])
def api_registrar():
    datos      = request.get_json(silent=True) or {}
    nombre     = (datos.get("name")     or "").strip()
    apellido   = (datos.get("lastname") or "").strip()
    correo     = (datos.get("email")    or "").strip().lower()
    contrasena =  datos.get("password") or ""

    if not all([nombre, apellido, correo, contrasena]):
        return jsonify({"success": False, "message": "Todos los campos son requeridos"}), 400
    if len(contrasena) < 8:
        return jsonify({"success": False, "message": "La contraseña debe tener al menos 8 caracteres"}), 400
    if not _validar_correo(correo):
        return jsonify({"success": False, "message": "Correo electrónico inválido"}), 400

    usuario = _crear_usuario(nombre, apellido, correo, contrasena)
    if usuario is None:
        return jsonify({"success": False, "message": "El correo ya está registrado"}), 409

    return jsonify({"success": True, "message": "Cuenta creada exitosamente."}), 201


@app.route("/api/logout", methods=["POST"])
def api_cerrar_sesion():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/me", methods=["GET"])
def api_mi_info():
    nom = session.get("nombreUsuario")
    if not nom:
        return jsonify({"success": False, "message": "No autenticado"}), 401

    usuario = _buscar_usuario(nom)
    if not usuario:
        return jsonify({"success": False, "message": "Usuario no encontrado"}), 401

    return jsonify({
        "success": True,
        "user": {
            "id":       usuario["id"],
            "name":     usuario["name"],
            "lastname": usuario["lastname"],
            "email":    usuario["email"],
            "role":     usuario["role"],
            "telefono": usuario.get("telefono") or "",
            "cargo":    usuario.get("cargo")    or "",
            "empresa":  usuario.get("empresa")  or "FIDUCCI",
        }
    })


@app.route("/api/forgot-password", methods=["POST"])
def api_recuperar_contrasena():
    datos  = request.get_json(silent=True) or {}
    correo = (datos.get("email") or "").strip().lower()
    if not correo:
        return jsonify({"success": False, "message": "Correo requerido"}), 400
    return jsonify({
        "success": True,
        "message": "Si el correo existe, recibirás instrucciones para restablecer tu contraseña."
    })


# Pedidos del perfil de usuario.

@app.route("/api/perfil", methods=["PUT"])
def api_actualizar_perfil():
    error = _verificar_auth()
    if error:
        return error

    id_usuario = session.get("idUsuario")
    datos      = request.get_json(silent=True) or {}

    nombre   = (datos.get("nombre")   or "").strip()
    apellido = (datos.get("apellido") or "").strip()
    correo   = (datos.get("email")    or "").strip().lower()

    if not nombre:
        return jsonify({"success": False, "mensaje": "El nombre es requerido"}), 400
    if not apellido:
        return jsonify({"success": False, "mensaje": "El apellido es requerido"}), 400
    if not correo or not _validar_correo(correo):
        return jsonify({"success": False, "mensaje": "Correo inválido"}), 400

    cambios = {
        "name":     nombre,
        "lastname": apellido,
        "email":    correo,
        "telefono": (datos.get("telefono") or "").strip(),
        "cargo":    (datos.get("cargo")    or "").strip(),
        "empresa":  (datos.get("empresa")  or "").strip(),
    }

    if datos.get("pw_nueva"):
        pw_actual  = datos.get("pw_actual",  "")
        pw_nueva   = datos.get("pw_nueva",   "")
        pw_confirm = datos.get("pw_confirm", "")

        usuario = _buscar_usuario(session.get("nombreUsuario"))
        if not usuario or not check_password_hash(usuario["password_hash"], pw_actual):
            return jsonify({"success": False, "mensaje": "Contraseña actual incorrecta"}), 400
        if len(pw_nueva) < 8:
            return jsonify({"success": False, "mensaje": "La nueva contraseña debe tener al menos 8 caracteres"}), 400
        if pw_nueva != pw_confirm:
            return jsonify({"success": False, "mensaje": "Las contraseñas no coinciden"}), 400

        cambios["password_hash"] = generate_password_hash(pw_nueva)

    actualizado = _actualizar_usuario(id_usuario, cambios)
    if not actualizado:
        return jsonify({"success": False, "mensaje": "Error al actualizar perfil"}), 500

    session["nombreUsuario"] = actualizado.get("username", session.get("nombreUsuario"))

    return jsonify({
        "success": True,
        "user": {
            "id":       actualizado["id"],
            "name":     actualizado["name"],
            "lastname": actualizado["lastname"],
            "email":    actualizado["email"],
            "role":     actualizado["role"],
            "telefono": actualizado.get("telefono") or "",
            "cargo":    actualizado.get("cargo")    or "",
            "empresa":  actualizado.get("empresa")  or "FIDUCCI",
        }
    })


# Pedidos de contactos.

@app.route("/api/contactos", methods=["GET"])
def api_get_contactos():
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    return jsonify({"success": True, "contactos": _obtener_contactos(id_usuario)})


@app.route("/api/contactos/nuevo", methods=["POST"])
def api_crear_contacto():
    error = _verificar_auth()
    if error:
        return error

    id_usuario = session.get("idUsuario")
    datos      = request.get_json(silent=True) or {}

    error_validacion = _validar_contacto_api(datos)
    if error_validacion:
        return jsonify({"success": False, "mensaje": error_validacion}), 400

    contacto = _crear_contacto(datos, id_usuario)
    return jsonify({"success": True, "contacto": contacto}), 201


@app.route("/api/contactos/<id_contacto>", methods=["GET"])
def api_get_contacto(id_contacto):
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    contactos  = _obtener_contactos(id_usuario)
    c = next((x for x in contactos if x["id"] == id_contacto), None)
    if not c:
        return jsonify({"success": False, "mensaje": "Contacto no encontrado"}), 404
    return jsonify({"success": True, "contacto": c})


@app.route("/api/contactos/<id_contacto>", methods=["PUT"])
def api_editar_contacto(id_contacto):
    error = _verificar_auth()
    if error:
        return error

    id_usuario = session.get("idUsuario")
    datos      = request.get_json(silent=True) or {}

    error_validacion = _validar_contacto_api(datos)
    if error_validacion:
        return jsonify({"success": False, "mensaje": error_validacion}), 400

    contacto = _actualizar_contacto(id_contacto, datos, id_usuario)
    if contacto is None:
        return jsonify({"success": False, "mensaje": "Contacto no encontrado"}), 404
    return jsonify({"success": True, "contacto": contacto})


@app.route("/api/contactos/<id_contacto>", methods=["DELETE"])
def api_eliminar_contacto(id_contacto):
    error = _verificar_auth()
    if error:
        return error

    id_usuario = session.get("idUsuario")
    if not _eliminar_contacto(id_contacto, id_usuario):
        return jsonify({"success": False, "mensaje": "Contacto no encontrado"}), 404
    return jsonify({"success": True})


@app.route("/api/contactos/<id_contacto>/favorito", methods=["POST"])
def api_cambiar_favorito(id_contacto):
    error = _verificar_auth()
    if error:
        return error

    id_usuario  = session.get("idUsuario")
    nuevo_estado = _alternar_favorito(id_contacto, id_usuario)
    if nuevo_estado is None:
        return jsonify({"success": False, "mensaje": "Contacto no encontrado"}), 404
    return jsonify({"success": True, "es_favorito": nuevo_estado})


# Pedidos de actividad.

@app.route("/api/actividad", methods=["GET"])
def api_get_actividad():
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    return jsonify({"success": True, "actividad": _obtener_actividad(id_usuario)})


# Pedido del reporte.

@app.route("/api/reporte", methods=["GET"])
def api_reporte():
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    reporte    = _generar_reporte(id_usuario)
    return jsonify({"success": True, "reporte": reporte})


# Pedido para importar contactos.

@app.route("/api/contactos/importar", methods=["POST"])
def api_importar_contactos():
    error = _verificar_auth()
    if error:
        return error

    id_usuario = session.get("idUsuario")
    archivo    = request.files.get("archivo")
    if not archivo:
        return jsonify({"success": False, "mensaje": "Archivo requerido"}), 400

    try:
        mapeo = json.loads(request.form.get("mapeo", "{}"))
    except Exception:
        mapeo = {}

    nombre_archivo = archivo.filename.lower()
    importados     = 0
    errores        = 0

    try:
        if nombre_archivo.endswith(".csv"):
            contenido = archivo.read().decode("utf-8-sig")
            lector    = modulo_csv.DictReader(io.StringIO(contenido))
            for fila in lector:
                datos = {campo: (fila.get(col) or "").strip()
                         for campo, col in mapeo.items() if col}
                if not datos.get("categoria"):
                    datos["categoria"] = "Otro"
                if not _validar_contacto_api(datos):
                    try:
                        _crear_contacto(datos, id_usuario)
                        importados += 1
                    except Exception:
                        errores += 1
                else:
                    errores += 1

        elif nombre_archivo.endswith((".xlsx", ".xls")):
            libro_imp  = openpyxl.load_workbook(io.BytesIO(archivo.read()))
            hoja_imp   = libro_imp.active
            encabezados = [str(hoja_imp.cell(1, j).value or "")
                           for j in range(1, hoja_imp.max_column + 1)]
            for i in range(2, hoja_imp.max_row + 1):
                fila  = {encabezados[j - 1]: str(hoja_imp.cell(i, j).value or "").strip()
                         for j in range(1, hoja_imp.max_column + 1)}
                datos = {campo: fila.get(col, "")
                         for campo, col in mapeo.items() if col}
                if not datos.get("categoria"):
                    datos["categoria"] = "Otro"
                if not _validar_contacto_api(datos):
                    try:
                        _crear_contacto(datos, id_usuario)
                        importados += 1
                    except Exception:
                        errores += 1
                else:
                    errores += 1
        else:
            return jsonify({"success": False, "mensaje": "Formato no soportado"}), 400

    except Exception as e:
        return jsonify({"success": False, "mensaje": f"Error al procesar archivo: {e}"}), 500

    return jsonify({
        "success":    True,
        "importados": importados,
        "errores":    errores,
        "mensaje":    f"Se importaron {importados} contactos."
    })


# Pedidos de conversaciones y notas.

@app.route("/api/contactos/<id_contacto>/interacciones", methods=["GET"])
def api_get_interacciones(id_contacto):
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    return jsonify({"success": True, "interacciones": _obtener_interacciones(id_contacto, id_usuario)})


@app.route("/api/contactos/<id_contacto>/interacciones", methods=["POST"])
def api_crear_interaccion(id_contacto):
    error = _verificar_auth()
    if error:
        return error

    id_usuario    = session.get("idUsuario")
    usuario       = _buscar_usuario(session.get("nombreUsuario", ""))
    nombre_usu    = f"{usuario.get('name', '')} {usuario.get('lastname', '')}".strip() if usuario else ""

    datos       = request.get_json(silent=True) or {}
    tipo        = (datos.get("tipo")        or "nota").strip()
    descripcion = (datos.get("descripcion") or "").strip()

    if not descripcion:
        return jsonify({"success": False, "mensaje": "La descripción es requerida"}), 400

    interaccion = _crear_interaccion(id_contacto, tipo, descripcion, nombre_usu, id_usuario)
    return jsonify({"success": True, "interaccion": interaccion}), 201


@app.route("/api/interacciones/<id_interaccion>", methods=["DELETE"])
def api_eliminar_interaccion(id_interaccion):
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    if not _eliminar_interaccion(id_interaccion, id_usuario):
        return jsonify({"success": False, "mensaje": "Interacción no encontrada"}), 404
    return jsonify({"success": True})


# Pedidos de recordatorios.

@app.route("/api/recordatorios", methods=["GET"])
def api_get_recordatorios():
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    return jsonify({"success": True, "recordatorios": _obtener_recordatorios(id_usuario)})


@app.route("/api/contactos/<id_contacto>/recordatorios", methods=["GET"])
def api_get_recordatorios_contacto(id_contacto):
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    return jsonify({"success": True, "recordatorios": _obtener_recordatorios_contacto(id_contacto, id_usuario)})


@app.route("/api/contactos/<id_contacto>/recordatorios", methods=["POST"])
def api_crear_recordatorio(id_contacto):
    error = _verificar_auth()
    if error:
        return error

    id_usuario  = session.get("idUsuario")
    datos       = request.get_json(silent=True) or {}
    titulo      = (datos.get("titulo")       or "").strip()
    fecha_limite = (datos.get("fecha_limite") or "").strip()

    if not titulo:
        return jsonify({"success": False, "mensaje": "El título es requerido"}), 400
    if not fecha_limite:
        return jsonify({"success": False, "mensaje": "La fecha es requerida"}), 400

    rec = _crear_recordatorio(id_contacto, titulo, fecha_limite, id_usuario)
    return jsonify({"success": True, "recordatorio": rec}), 201


@app.route("/api/recordatorios/<id_rec>/completar", methods=["POST"])
def api_completar_recordatorio(id_rec):
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    if _completar_recordatorio(id_rec, id_usuario) is None:
        return jsonify({"success": False, "mensaje": "Recordatorio no encontrado"}), 404
    return jsonify({"success": True})


@app.route("/api/recordatorios/<id_rec>", methods=["DELETE"])
def api_eliminar_recordatorio(id_rec):
    error = _verificar_auth()
    if error:
        return error
    id_usuario = session.get("idUsuario")
    if not _eliminar_recordatorio(id_rec, id_usuario):
        return jsonify({"success": False, "mensaje": "Recordatorio no encontrado"}), 404
    return jsonify({"success": True})


# Errores.

@app.errorhandler(404)
def no_encontrado(_):
    return jsonify({"success": False, "message": "Recurso no encontrado"}), 404


@app.errorhandler(500)
def error_servidor(e):
    print(f"Error interno: {e}")
    return jsonify({"success": False, "message": "Error interno del servidor"}), 500


# Inicio del programa.

if __name__ == "__main__":
    _cargar_libro()
    print("FIDUCCI - Servidor activo en http://localhost:5000")
    print("Usuario: admin  |  Contraseña: 1234")
    app.run(
        debug=os.environ.get("FLASK_DEBUG", "0") == "1",
        port=int(os.environ.get("PORT", "5000")),
        use_reloader=False,
    )
